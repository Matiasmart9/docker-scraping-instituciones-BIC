import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_excel_instituciones(datos_instituciones: list) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) estilizado profesionalmente con la información
    y el estado auditado de las instituciones financieras.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Instituciones BICSA"

    # Habilitar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # Paleta de Colores
    NAVY_HEADER = "1E293B"       # Slate Dark
    WHITE_TEXT = "FFFFFF"
    ZEBRA_FILL = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    
    # Estilos de Alerta
    FILL_CRITICO = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red light
    FONT_CRITICO = Font(name="Calibri", size=10, bold=True, color="991B1B")

    FILL_ADVERTENCIA = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Amber light
    FONT_ADVERTENCIA = Font(name="Calibri", size=10, bold=True, color="92400E")

    FILL_BLOQUEADO = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    FONT_BLOQUEADO = Font(name="Calibri", size=10, bold=True, color="475569")

    FILL_NORMAL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Green light
    FONT_NORMAL = Font(name="Calibri", size=10, bold=True, color="166534")

    # Fuentes Estándar
    font_title = Font(name="Calibri", size=16, bold=True, color="0F172A")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="64748B")
    font_header = Font(name="Calibri", size=11, bold=True, color=WHITE_TEXT)
    font_cell = Font(name="Calibri", size=10, color="0F172A")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")

    # Bordes
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # 1. Título y Encabezado del Reporte
    ws.merge_cells("A1:J1")
    ws["A1"] = "MONITOREO DE ESTADO DE INSTITUCIONES - BICSA"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Reporte generado automáticamente el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = Alignment(vertical="center")

    # 2. Resumen de KPIs
    total_inst = len(datos_instituciones)
    activas = sum(1 for d in datos_instituciones if "ACTIV" in str(d.get("estado", "")).upper())
    bloqueadas = sum(1 for d in datos_instituciones if "BLOQUEA" in str(d.get("estado", "")).upper())
    criticas = sum(1 for d in datos_instituciones if d.get("nivel_alerta") == "CRITICO")

    ws["A4"] = "Total Registros"
    ws["B4"] = total_inst
    ws["C4"] = "Activas"
    ws["D4"] = activas
    ws["E4"] = "Bloqueadas"
    ws["F4"] = bloqueadas
    ws["G4"] = "Alertas Críticas"
    ws["H4"] = criticas

    for col in ["A", "C", "E", "G"]:
        ws[f"{col}4"].font = font_bold
        ws[f"{col}4"].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    for col in ["B", "D", "F", "H"]:
        ws[f"{col}4"].font = font_bold
        ws[f"{col}4"].alignment = Alignment(horizontal="center")

    # 3. Encabezados de Tabla
    headers = [
        "Nombre Institución",
        "Estado Actual",
        "Categoría Tabla",
        "Búsquedas Máx.",
        "Fecha Última Carga",
        "Calidad Datos",
        "Motivo Suspensión",
        "Vencimiento Validación",
        "Horas Hábiles Restantes",
        "Nivel Alerta"
    ]

    row_idx = 6
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[row_idx].height = 28

    # 4. Filas de Datos
    row_idx = 7
    for item in datos_instituciones:
        ws.cell(row=row_idx, column=1, value=item.get("nombre", "")).font = font_bold
        ws.cell(row=row_idx, column=2, value=item.get("estado", "")).font = font_cell
        ws.cell(row=row_idx, column=3, value=item.get("categoria_tabla", "")).font = font_cell
        
        c_busq = ws.cell(row=row_idx, column=4, value=item.get("cant_max_busquedas", 0))
        c_busq.font = font_cell
        c_busq.number_format = '#,##0'
        c_busq.alignment = Alignment(horizontal="right")

        raw_fecha = str(item.get("fecha_ultima_carga", "N/A") or "N/A").strip()
        clean_fecha = raw_fecha.split(" ")[0] if raw_fecha and raw_fecha not in ["None", "N/A", "N.A"] else "N/A"
        ws.cell(row=row_idx, column=5, value=clean_fecha).font = font_cell
        ws.cell(row=row_idx, column=6, value=item.get("calidad_datos", "N.A")).font = font_cell
        ws.cell(row=row_idx, column=7, value=item.get("motivo_suspension", "-")).font = font_cell
        ws.cell(row=row_idx, column=8, value=item.get("vencimiento_validacion", "-")).font = font_cell
        
        c_hrs = ws.cell(row=row_idx, column=9, value=item.get("horas_habiles_restantes", 72.0))
        c_hrs.font = font_cell
        c_hrs.number_format = '0.0 "hs"'
        c_hrs.alignment = Alignment(horizontal="right")

        # Nivel de Alerta con estilo condicional
        alerta_cell = ws.cell(row=row_idx, column=10, value=item.get("nivel_alerta", "NORMAL"))
        alerta_val = str(item.get("nivel_alerta", "NORMAL")).upper()
        alerta_cell.alignment = Alignment(horizontal="center")

        if alerta_val == "CRITICO":
            alerta_cell.fill = FILL_CRITICO
            alerta_cell.font = FONT_CRITICO
        elif alerta_val == "ADVERTENCIA":
            alerta_cell.fill = FILL_ADVERTENCIA
            alerta_cell.font = FONT_ADVERTENCIA
        elif alerta_val == "BLOQUEADO":
            alerta_cell.fill = FILL_BLOQUEADO
            alerta_cell.font = FONT_BLOQUEADO
        else:
            alerta_cell.fill = FILL_NORMAL
            alerta_cell.font = FONT_NORMAL

        # Aplicar bordes a la fila
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

        row_idx += 1

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions["A"].width = 45 # Nombre de Institución más amplio

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
